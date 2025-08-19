import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import math

# Create workbook with multiple sheets
wb = Workbook()

# Remove default sheet
wb.remove(wb.active)

# Create sheets
shallow_sheet = wb.create_sheet('Shallow Foundation')
pile_sheet = wb.create_sheet('Pile Design')

# Define styles
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
input_fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
calc_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
result_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
bold_font = Font(bold=True)

# SHALLOW FOUNDATION SHEET
ws1 = shallow_sheet

# Title
ws1['A1'] = 'SHALLOW FOUNDATION BEARING CAPACITY DESIGN'
ws1['A1'].font = Font(bold=True, size=16)
ws1.merge_cells('A1:F1')

# Project Information
row = 3
ws1[f'A{row}'] = 'PROJECT INFORMATION'
ws1[f'A{row}'].font = header_font
ws1[f'A{row}'].fill = header_fill
ws1.merge_cells(f'A{row}:B{row}')

row += 1
ws1[f'A{row}'] = 'Project:'
ws1[f'B{row}'] = '[Enter Project Name]'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Engineer:'
ws1[f'B{row}'] = '[Enter Engineer Name]'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Date:'
ws1[f'B{row}'] = '[Enter Date]'
ws1[f'B{row}'].fill = input_fill

# INPUT PARAMETERS
row += 3
ws1[f'A{row}'] = 'INPUT PARAMETERS'
ws1[f'A{row}'].font = header_font
ws1[f'A{row}'].fill = header_fill
ws1.merge_cells(f'A{row}:D{row}')

row += 2
# Soil Properties
ws1[f'A{row}'] = 'SOIL PROPERTIES'
ws1[f'A{row}'].font = bold_font
row += 1
ws1[f'A{row}'] = 'Unit Weight (γ):'
ws1[f'B{row}'] = 115
ws1[f'C{row}'] = 'pcf'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Effective Friction Angle (φ):'
ws1[f'B{row}'] = 30
ws1[f'C{row}'] = 'degrees'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Cohesion (c):'
ws1[f'B{row}'] = 0
ws1[f'C{row}'] = 'psf'
ws1[f'B{row}'].fill = input_fill

row += 2
# Foundation Properties
ws1[f'A{row}'] = 'FOUNDATION PROPERTIES'
ws1[f'A{row}'].font = bold_font
row += 1
ws1[f'A{row}'] = 'Width (B):'
ws1[f'B{row}'] = 6.5
ws1[f'C{row}'] = 'ft'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Length (L):'
ws1[f'B{row}'] = 10.0
ws1[f'C{row}'] = 'ft'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Depth of Foundation (Df):'
ws1[f'B{row}'] = 5.0
ws1[f'C{row}'] = 'ft'
ws1[f'B{row}'].fill = input_fill

row += 2
# Load Properties
ws1[f'A{row}'] = 'LOAD PROPERTIES'
ws1[f'A{row}'].font = bold_font
row += 1
ws1[f'A{row}'] = 'Vertical Load (V):'
ws1[f'B{row}'] = 225000
ws1[f'C{row}'] = 'lbs'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Horizontal Load (H):'
ws1[f'B{row}'] = 11250
ws1[f'C{row}'] = 'lbs'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Moment (M):'
ws1[f'B{row}'] = 55300
ws1[f'C{row}'] = 'lb·ft'
ws1[f'B{row}'].fill = input_fill

row += 2
# Water Table
ws1[f'A{row}'] = 'WATER TABLE'
ws1[f'A{row}'].font = bold_font
row += 1
ws1[f'A{row}'] = 'Depth to Water Table:'
ws1[f'B{row}'] = 16.5
ws1[f'C{row}'] = 'ft'
ws1[f'B{row}'].fill = input_fill

# CALCULATIONS SECTION
row += 3
ws1[f'A{row}'] = 'CALCULATIONS'
ws1[f'A{row}'].font = header_font
ws1[f'A{row}'].fill = header_fill
ws1.merge_cells(f'A{row}:E{row}')

row += 2
# Bearing Capacity Factors
ws1[f'A{row}'] = 'BEARING CAPACITY FACTORS'
ws1[f'A{row}'].font = bold_font
row += 1
ws1[f'A{row}'] = 'Nc = (Nq - 1) * cot(φ)'
ws1[f'B{row}'] = '=((EXP(PI()*(TAN(RADIANS(B14))))*((TAN(RADIANS(45+B14/2)))^2))-1)*1/TAN(RADIANS(B14))'
ws1[f'C{row}'] = 'Terzaghi/Meyerhof'
ws1[f'B{row}'].fill = calc_fill
row += 1
ws1[f'A{row}'] = 'Nq = e^(π*tan φ) * tan²(45° + φ/2)'
ws1[f'B{row}'] = '=EXP(PI()*TAN(RADIANS(B14)))*(TAN(RADIANS(45+B14/2)))^2'
ws1[f'C{row}'] = 'Terzaghi/Meyerhof'
ws1[f'B{row}'].fill = calc_fill
row += 1
ws1[f'A{row}'] = 'Nγ = 2 * (Nq + 1) * tan φ'
ws1[f'B{row}'] = '=2*(B28+1)*TAN(RADIANS(B14))'
ws1[f'C{row}'] = 'Meyerhof'
ws1[f'B{row}'].fill = calc_fill

row += 2
# Shape Factors
ws1[f'A{row}'] = 'SHAPE FACTORS (De Beer/Vesić)'
ws1[f'A{row}'].font = bold_font
row += 1
ws1[f'A{row}'] = 'sc = 1 + (B/L) * (Nq/Nc)'
ws1[f'B{row}'] = '=1+(B17/B18)*(B28/B27)'
ws1[f'B{row}'].fill = calc_fill
row += 1
ws1[f'A{row}'] = 'sq = 1 + (B/L) * tan φ'
ws1[f'B{row}'] = '=1+(B17/B18)*TAN(RADIANS(B14))'
ws1[f'B{row}'].fill = calc_fill
row += 1
ws1[f'A{row}'] = 'sγ = 1 - 0.4 * (B/L)'
ws1[f'B{row}'] = '=1-0.4*(B17/B18)'
ws1[f'B{row}'].fill = calc_fill

row += 2
# Depth Factors
ws1[f'A{row}'] = 'DEPTH FACTORS'
ws1[f'A{row}'].font = bold_font
row += 1
ws1[f'A{row}'] = 'dc = 1 + 0.4 * (Df/B)'
ws1[f'B{row}'] = '=1+0.4*(B19/B17)'
ws1[f'B{row}'].fill = calc_fill
row += 1
ws1[f'A{row}'] = 'dq = 1 + 2 * tan φ * (1 - sin φ)² * (Df/B)'
ws1[f'B{row}'] = '=1+2*TAN(RADIANS(B14))*(1-SIN(RADIANS(B14)))^2*(B19/B17)'
ws1[f'B{row}'].fill = calc_fill
row += 1
ws1[f'A{row}'] = 'dγ = 1.0'
ws1[f'B{row}'] = 1.0
ws1[f'B{row}'].fill = calc_fill

row += 2
# Ultimate Bearing Capacity
ws1[f'A{row}'] = 'ULTIMATE BEARING CAPACITY'
ws1[f'A{row}'].font = bold_font
row += 1
ws1[f'A{row}'] = 'qu = c*Nc*sc*dc + q*Nq*sq*dq + 0.5*γ*B*Nγ*sγ*dγ'
ws1[f'A{row}'].fill = result_fill
row += 1
ws1[f'A{row}'] = 'Cohesion Term:'
ws1[f'B{row}'] = '=B15*B27*B31*B35'
ws1[f'C{row}'] = 'psf'
ws1[f'B{row}'].fill = calc_fill
row += 1
ws1[f'A{row}'] = 'Surcharge Term:'
ws1[f'B{row}'] = '=(B12*B19)*(B28*B32*B36)'
ws1[f'C{row}'] = 'psf'
ws1[f'B{row}'].fill = calc_fill
row += 1
ws1[f'A{row}'] = 'Weight Term:'
ws1[f'B{row}'] = '=0.5*B12*B17*B29*B33*B37'
ws1[f'C{row}'] = 'psf'
ws1[f'B{row}'].fill = calc_fill
row += 1
ws1[f'A{row}'] = 'qu (Ultimate):'
ws1[f'B{row}'] = '=B41+B42+B43'
ws1[f'C{row}'] = 'psf'
ws1[f'B{row}'].fill = result_fill
ws1[f'B{row}'].font = bold_font

row += 2
# Allowable Bearing Capacity
ws1[f'A{row}'] = 'ALLOWABLE BEARING CAPACITY'
ws1[f'A{row}'].font = bold_font
row += 1
ws1[f'A{row}'] = 'Factor of Safety:'
ws1[f'B{row}'] = 3.0
ws1[f'C{row}'] = '(FHWA Recommended: 2.5-3.0)'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'qa (Allowable):'
ws1[f'B{row}'] = '=B44/B46'
ws1[f'C{row}'] = 'psf'
ws1[f'B{row}'].fill = result_fill
ws1[f'B{row}'].font = bold_font

row += 2
# Applied Pressure Check
ws1[f'A{row}'] = 'APPLIED PRESSURE CHECK'
ws1[f'A{row}'].font = bold_font
row += 1
ws1[f'A{row}'] = 'Applied Pressure:'
ws1[f'B{row}'] = '=B20/(B17*B18)'
ws1[f'C{row}'] = 'psf'
ws1[f'B{row}'].fill = calc_fill
row += 1
ws1[f'A{row}'] = 'Safety Factor:'
ws1[f'B{row}'] = '=B47/B49'
ws1[f'C{row}'] = '(Must be > 2.5)'
ws1[f'B{row}'].fill = result_fill
row += 1
ws1[f'A{row}'] = 'Status:'
ws1[f'B{row}'] = '=IF(B50>2.5,"SAFE","UNSAFE")'
ws1[f'B{row}'].fill = result_fill
ws1[f'B{row}'].font = bold_font

# PILE DESIGN SHEET
ws2 = pile_sheet

# Title
ws2['A1'] = 'PILE DESIGN PER FHWA GUIDELINES'
ws2['A1'].font = Font(bold=True, size=16)
ws2.merge_cells('A1:F1')

# Project Information
row = 3
ws2[f'A{row}'] = 'PROJECT INFORMATION'
ws2[f'A{row}'].font = header_font
ws2[f'A{row}'].fill = header_fill
ws2.merge_cells(f'A{row}:B{row}')

row += 1
ws2[f'A{row}'] = 'Project:'
ws2[f'B{row}'] = '[Enter Project Name]'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Engineer:'
ws2[f'B{row}'] = '[Enter Engineer Name]'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Date:'
ws2[f'B{row}'] = '[Enter Date]'
ws2[f'B{row}'].fill = input_fill

# INPUT PARAMETERS
row += 3
ws2[f'A{row}'] = 'INPUT PARAMETERS'
ws2[f'A{row}'].font = header_font
ws2[f'A{row}'].fill = header_fill
ws2.merge_cells(f'A{row}:D{row}')

row += 2
# Pile Properties
ws2[f'A{row}'] = 'PILE PROPERTIES'
ws2[f'A{row}'].font = bold_font
row += 1
ws2[f'A{row}'] = 'Pile Type:'
ws2[f'B{row}'] = 'Driven Steel H-Pile'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Pile Diameter/Width (B):'
ws2[f'B{row}'] = 14
ws2[f'C{row}'] = 'in'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Pile Length (L):'
ws2[f'B{row}'] = 50
ws2[f'C{row}'] = 'ft'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Perimeter (P):'
ws2[f'B{row}'] = '=PI()*B15/12'
ws2[f'C{row}'] = 'ft'
ws2[f'B{row}'].fill = calc_fill

row += 2
# Soil Layers (simplified - single layer)
ws2[f'A{row}'] = 'SOIL PROPERTIES'
ws2[f'A{row}'].font = bold_font
row += 1
ws2[f'A{row}'] = 'Unit Weight (γ):'
ws2[f'B{row}'] = 115
ws2[f'C{row}'] = 'pcf'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Effective Friction Angle (φ):'
ws2[f'B{row}'] = 30
ws2[f'C{row}'] = 'degrees'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Cohesion (c):'
ws2[f'B{row}'] = 520
ws2[f'C{row}'] = 'psf'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'SPT N-Value:'
ws2[f'B{row}'] = 20
ws2[f'C{row}'] = 'blows/12in'
ws2[f'B{row}'].fill = input_fill

row += 2
# Load Properties
ws2[f'A{row}'] = 'LOAD PROPERTIES'
ws2[f'A{row}'].font = bold_font
row += 1
ws2[f'A{row}'] = 'Design Load (P):'
ws2[f'B{row}'] = 112500
ws2[f'C{row}'] = 'lbs'
ws2[f'B{row}'].fill = input_fill

# CALCULATIONS SECTION
row += 3
ws2[f'A{row}'] = 'CALCULATIONS'
ws2[f'A{row}'].font = header_font
ws2[f'A{row}'].fill = header_fill
ws2.merge_cells(f'A{row}:E{row}')

row += 2
# Skin Friction
ws2[f'A{row}'] = 'SKIN FRICTION CAPACITY'
ws2[f'A{row}'].font = bold_font
row += 1
ws2[f'A{row}'] = 'Alpha Method (Clay):'
ws2[f'A{row}'].font = bold_font
row += 1
ws2[f'A{row}'] = 'Alpha Factor (α):'
ws2[f'B{row}'] = 0.5
ws2[f'C{row}'] = '(Conservative for driven piles)'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'fs (clay) = α * c:'
ws2[f'B{row}'] = '=B33*B22'
ws2[f'C{row}'] = 'psf'
ws2[f'B{row}'].fill = calc_fill
row += 1
ws2[f'A{row}'] = 'Qs (clay) = fs * P * L:'
ws2[f'B{row}'] = '=B34*B17*B16'
ws2[f'C{row}'] = 'lbs'
ws2[f'B{row}'].fill = calc_fill

row += 2
ws2[f'A{row}'] = 'Beta Method (Sand):'
ws2[f'A{row}'].font = bold_font
row += 1
ws2[f'A{row}'] = 'Beta Factor (β):'
ws2[f'B{row}'] = '=TAN(RADIANS(B21))*0.8'
ws2[f'C{row}'] = '(K * tan φ, K=0.8 for driven)'
ws2[f'B{row}'].fill = calc_fill
row += 1
ws2[f'A{row}'] = 'Effective Stress at Mid-depth:'
ws2[f'B{row}'] = '=B19*B16/2'
ws2[f'C{row}'] = 'psf'
ws2[f'B{row}'].fill = calc_fill
row += 1
ws2[f'A{row}'] = 'fs (sand) = β * σ_v:'
ws2[f'B{row}'] = '=B37*B38'
ws2[f'C{row}'] = 'psf'
ws2[f'B{row}'].fill = calc_fill
row += 1
ws2[f'A{row}'] = 'Qs (sand) = fs * P * L:'
ws2[f'B{row}'] = '=B39*B17*B16'
ws2[f'C{row}'] = 'lbs'
ws2[f'B{row}'].fill = calc_fill

row += 2
# End Bearing
ws2[f'A{row}'] = 'END BEARING CAPACITY'
ws2[f'A{row}'].font = bold_font
row += 1
ws2[f'A{row}'] = 'Nq (for piles):'
ws2[f'B{row}'] = '=EXP(PI()*TAN(RADIANS(B21)))*(TAN(RADIANS(45+B21/2)))^2'
ws2[f'C{row}'] = 'Bearing capacity factor'
ws2[f'B{row}'].fill = calc_fill
row += 1
ws2[f'A{row}'] = 'Effective Stress at Tip:'
ws2[f'B{row}'] = '=B19*B16'
ws2[f'C{row}'] = 'psf'
ws2[f'B{row}'].fill = calc_fill
row += 1
ws2[f'A{row}'] = 'qp = c*Nc + σ_v*Nq (simplified):'
ws2[f'B{row}'] = '=B22*9.14+B43*B42'
ws2[f'C{row}'] = 'psf (Nc=9.14 for φ=30°)'
ws2[f'B{row}'].fill = calc_fill
row += 1
ws2[f'A{row}'] = 'Qp = qp * Ap:'
ws2[f'B{row}'] = '=B44*PI()*((B15/12/2)^2)'
ws2[f'C{row}'] = 'lbs'
ws2[f'B{row}'].fill = calc_fill

row += 2
# Ultimate Capacity
ws2[f'A{row}'] = 'ULTIMATE PILE CAPACITY'
ws2[f'A{row}'].font = bold_font
row += 1
ws2[f'A{row}'] = 'Qu = Qs + Qp:'
ws2[f'B{row}'] = '=MAX(B35,B40)+B45'
ws2[f'C{row}'] = 'lbs (use max of clay or sand method)'
ws2[f'B{row}'].fill = result_fill
ws2[f'B{row}'].font = bold_font

row += 2
# Allowable Capacity
ws2[f'A{row}'] = 'ALLOWABLE PILE CAPACITY'
ws2[f'A{row}'].font = bold_font
row += 1
ws2[f'A{row}'] = 'Factor of Safety:'
ws2[f'B{row}'] = 2.5
ws2[f'C{row}'] = '(FHWA Recommended: 2.0-2.5)'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Qa = Qu / FS:'
ws2[f'B{row}'] = '=B47/B49'
ws2[f'C{row}'] = 'lbs'
ws2[f'B{row}'].fill = result_fill
ws2[f'B{row}'].font = bold_font

row += 2
# Design Check
ws2[f'A{row}'] = 'DESIGN CHECK'
ws2[f'A{row}'].font = bold_font
row += 1
ws2[f'A{row}'] = 'Applied Load:'
ws2[f'B{row}'] = '=B25'
ws2[f'C{row}'] = 'lbs'
ws2[f'B{row}'].fill = calc_fill
row += 1
ws2[f'A{row}'] = 'Safety Factor:'
ws2[f'B{row}'] = '=B50/B52'
ws2[f'C{row}'] = '(Must be > 2.0)'
ws2[f'B{row}'].fill = result_fill
row += 1
ws2[f'A{row}'] = 'Status:'
ws2[f'B{row}'] = '=IF(B53>2.0,"SAFE","UNSAFE")'
ws2[f'B{row}'].fill = result_fill
ws2[f'B{row}'].font = bold_font

# Group Efficiency Section
row += 3
ws2[f'A{row}'] = 'PILE GROUP ANALYSIS'
ws2[f'A{row}'].font = header_font
ws2[f'A{row}'].fill = header_fill
ws2.merge_cells(f'A{row}:D{row}')

row += 2
ws2[f'A{row}'] = 'Number of Piles:'
ws2[f'B{row}'] = 4
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Pile Spacing:'
ws2[f'B{row}'] = 10.0
ws2[f'C{row}'] = 'ft (recommend 3D min)'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Group Efficiency:'
ws2[f'B{row}'] = 0.85
ws2[f'C{row}'] = '(Conservative estimate)'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Group Capacity:'
ws2[f'B{row}'] = '=B57*B50*B59'
ws2[f'C{row}'] = 'lbs'
ws2[f'B{row}'].fill = result_fill
ws2[f'B{row}'].font = bold_font

# Set column widths for both sheets
for sheet in [ws1, ws2]:
    sheet.column_dimensions['A'].width = 35
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15

# Save the workbook
wb.save('Foundation_Pile_Design_FHWA.xlsx')

print("Excel file 'Foundation_Pile_Design_FHWA.xlsx' created successfully!")
print("Features included:")
print("- Shallow Foundation Bearing Capacity (Terzaghi/Meyerhof/Vesic methods)")
print("- Pile Design with Skin Friction and End Bearing")
print("- FHWA recommended safety factors")
print("- Clear input/output sections with color coding")
print("- Step-by-step calculations with formulas")
print("- Pile group analysis")
print("- All units converted to English (psf, pcf, lbs, ft, in)")