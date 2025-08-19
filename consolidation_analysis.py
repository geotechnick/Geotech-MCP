import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import math

# Create workbook with multiple sheets
wb = Workbook()

# Remove default sheet
wb.remove(wb.active)

# Create sheets
immediate_sheet = wb.create_sheet('Immediate Settlement')
primary_sheet = wb.create_sheet('Primary Consolidation')
embankment_sheet = wb.create_sheet('Embankment Analysis')

# Define styles
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
input_fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
calc_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
result_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
bold_font = Font(bold=True)

# IMMEDIATE SETTLEMENT SHEET
ws1 = immediate_sheet

# Title
ws1['A1'] = 'IMMEDIATE SETTLEMENT ANALYSIS'
ws1['A1'].font = Font(bold=True, size=16)
ws1.merge_cells('A1:F1')

# Project Information
row = 3
ws1[f'A{row}'] = 'PROJECT INFORMATION'
ws1[f'A{row}'].font = header_font
ws1[f'A{row}'].fill = header_fill
ws1.merge_cells(f'A{row}:B{row}')

row += 1
ws1[f'A{row}'] = 'Project:'
ws1[f'B{row}'] = '[Enter Project Name]'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Engineer:'
ws1[f'B{row}'] = '[Enter Engineer Name]'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Date:'
ws1[f'B{row}'] = '[Enter Date]'
ws1[f'B{row}'].fill = input_fill

# INPUT PARAMETERS
row += 3
ws1[f'A{row}'] = 'INPUT PARAMETERS'
ws1[f'A{row}'].font = header_font
ws1[f'A{row}'].fill = header_fill
ws1.merge_cells(f'A{row}:D{row}')

row += 2
# Foundation Properties
ws1[f'A{row}'] = 'FOUNDATION PROPERTIES'
ws1[f'A{row}'].font = bold_font
row += 1
ws1[f'A{row}'] = 'Foundation Width (B):'
ws1[f'B{row}'] = 8.0
ws1[f'C{row}'] = 'ft'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Foundation Length (L):'
ws1[f'B{row}'] = 12.0
ws1[f'C{row}'] = 'ft'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Foundation Depth (Df):'
ws1[f'B{row}'] = 4.0
ws1[f'C{row}'] = 'ft'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Applied Pressure (q):'
ws1[f'B{row}'] = 2000
ws1[f'C{row}'] = 'psf'
ws1[f'B{row}'].fill = input_fill

row += 2
# Soil Properties
ws1[f'A{row}'] = 'SOIL PROPERTIES'
ws1[f'A{row}'].font = bold_font
row += 1
ws1[f'A{row}'] = 'Soil Type:'
ws1[f'B{row}'] = 'Clay'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Unit Weight (γ):'
ws1[f'B{row}'] = 120
ws1[f'C{row}'] = 'pcf'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Poisson Ratio (ν):'
ws1[f'B{row}'] = 0.35
ws1[f'C{row}'] = '(typical for clay)'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Modulus of Elasticity (E):'
ws1[f'B{row}'] = 5000
ws1[f'C{row}'] = 'psf'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Thickness of Layer (H):'
ws1[f'B{row}'] = 20
ws1[f'C{row}'] = 'ft'
ws1[f'B{row}'].fill = input_fill

row += 2
# Water Table
ws1[f'A{row}'] = 'WATER TABLE'
ws1[f'A{row}'].font = bold_font
row += 1
ws1[f'A{row}'] = 'Depth to Water Table:'
ws1[f'B{row}'] = 8.0
ws1[f'C{row}'] = 'ft'
ws1[f'B{row}'].fill = input_fill

# CALCULATIONS SECTION
row += 3
ws1[f'A{row}'] = 'IMMEDIATE SETTLEMENT CALCULATIONS'
ws1[f'A{row}'].font = header_font
ws1[f'A{row}'].fill = header_fill
ws1.merge_cells(f'A{row}:E{row}')

row += 2
# Shape and Depth Factors
ws1[f'A{row}'] = 'INFLUENCE FACTORS'
ws1[f'A{row}'].font = bold_font
row += 1
ws1[f'A{row}'] = 'Length/Width Ratio (L/B):'
ws1[f'B{row}'] = '=B13/B12'
ws1[f'B{row}'].fill = calc_fill
row += 1
ws1[f'A{row}'] = 'Shape Factor (Is):'
ws1[f'B{row}'] = '=1.12*SQRT((B13*B12)/(B13+B12))/SQRT(B12)'
ws1[f'C{row}'] = 'Janbu et al. (rectangular)'
ws1[f'B{row}'].fill = calc_fill
row += 1
ws1[f'A{row}'] = 'Depth/Width Ratio (Df/B):'
ws1[f'B{row}'] = '=B14/B12'
ws1[f'B{row}'].fill = calc_fill
row += 1
ws1[f'A{row}'] = 'Depth Factor (Id):'
ws1[f'B{row}'] = '=1-0.25*LOG10(B33+1)'
ws1[f'C{row}'] = 'Fox (1948)'
ws1[f'B{row}'].fill = calc_fill

row += 2
# Stress Distribution
ws1[f'A{row}'] = 'STRESS DISTRIBUTION'
ws1[f'A{row}'].font = bold_font
row += 1
ws1[f'A{row}'] = 'Net Applied Stress (Δq):'
ws1[f'B{row}'] = '=B15-B17*B14'
ws1[f'C{row}'] = 'psf (Applied - Removed Soil)'
ws1[f'B{row}'].fill = calc_fill
row += 1
ws1[f'A{row}'] = 'Depth to Center of Layer:'
ws1[f'B{row}'] = '=B14+B21/2'
ws1[f'C{row}'] = 'ft'
ws1[f'B{row}'].fill = calc_fill
row += 1
ws1[f'A{row}'] = 'Stress Influence Factor (I):'
ws1[f'B{row}'] = '=B32*B34'
ws1[f'C{row}'] = 'Combined Is * Id'
ws1[f'B{row}'].fill = calc_fill

row += 2
# Immediate Settlement Calculation
ws1[f'A{row}'] = 'IMMEDIATE SETTLEMENT (Elastic)'
ws1[f'A{row}'].font = bold_font
row += 1
ws1[f'A{row}'] = 'Si = q * B * (1-ν²) * I / E'
ws1[f'A{row}'].fill = result_fill
row += 1
ws1[f'A{row}'] = 'Immediate Settlement (Si):'
ws1[f'B{row}'] = '=B36*B12*(1-B19^2)*B38/B20'
ws1[f'C{row}'] = 'ft'
ws1[f'B{row}'].fill = result_fill
ws1[f'B{row}'].font = bold_font
row += 1
ws1[f'A{row}'] = 'Immediate Settlement (Si):'
ws1[f'B{row}'] = '=B41*12'
ws1[f'C{row}'] = 'inches'
ws1[f'B{row}'].fill = result_fill
ws1[f'B{row}'].font = bold_font

# Alternative Methods
row += 3
ws1[f'A{row}'] = 'ALTERNATIVE METHODS'
ws1[f'A{row}'].font = header_font
ws1[f'A{row}'].fill = header_fill
ws1.merge_cells(f'A{row}:E{row}')

row += 2
# Schmertmann Method
ws1[f'A{row}'] = 'SCHMERTMANN METHOD (SPT-Based)'
ws1[f'A{row}'].font = bold_font
row += 1
ws1[f'A{row}'] = 'SPT N-Value:'
ws1[f'B{row}'] = 15
ws1[f'C{row}'] = 'blows/12in'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Es = 2000 * (N + 15) for clay:'
ws1[f'B{row}'] = '=2000*(B47+15)'
ws1[f'C{row}'] = 'psf'
ws1[f'B{row}'].fill = calc_fill
row += 1
ws1[f'A{row}'] = 'Settlement (Schmertmann):'
ws1[f'B{row}'] = '=B36*B12*(1-B19^2)*B38/B48*12'
ws1[f'C{row}'] = 'inches'
ws1[f'B{row}'].fill = calc_fill

# PRIMARY CONSOLIDATION SHEET
ws2 = primary_sheet

# Title
ws2['A1'] = 'PRIMARY CONSOLIDATION ANALYSIS'
ws2['A1'].font = Font(bold=True, size=16)
ws2.merge_cells('A1:F1')

# Project Information
row = 3
ws2[f'A{row}'] = 'PROJECT INFORMATION'
ws2[f'A{row}'].font = header_font
ws2[f'A{row}'].fill = header_fill
ws2.merge_cells(f'A{row}:B{row}')

row += 1
ws2[f'A{row}'] = 'Project:'
ws2[f'B{row}'] = '[Enter Project Name]'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Engineer:'
ws2[f'B{row}'] = '[Enter Engineer Name]'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Date:'
ws2[f'B{row}'] = '[Enter Date]'
ws2[f'B{row}'].fill = input_fill

# INPUT PARAMETERS
row += 3
ws2[f'A{row}'] = 'INPUT PARAMETERS'
ws2[f'A{row}'].font = header_font
ws2[f'A{row}'].fill = header_fill
ws2.merge_cells(f'A{row}:D{row}')

row += 2
# Foundation Properties
ws2[f'A{row}'] = 'FOUNDATION PROPERTIES'
ws2[f'A{row}'].font = bold_font
row += 1
ws2[f'A{row}'] = 'Foundation Width (B):'
ws2[f'B{row}'] = 8.0
ws2[f'C{row}'] = 'ft'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Foundation Length (L):'
ws2[f'B{row}'] = 12.0
ws2[f'C{row}'] = 'ft'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Applied Pressure (Δq):'
ws2[f'B{row}'] = 2000
ws2[f'C{row}'] = 'psf'
ws2[f'B{row}'].fill = input_fill

row += 2
# Soil Properties
ws2[f'A{row}'] = 'CONSOLIDATING LAYER PROPERTIES'
ws2[f'A{row}'].font = bold_font
row += 1
ws2[f'A{row}'] = 'Layer Thickness (H):'
ws2[f'B{row}'] = 15
ws2[f'C{row}'] = 'ft'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Unit Weight (γ):'
ws2[f'B{row}'] = 120
ws2[f'C{row}'] = 'pcf'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Depth to Layer Top:'
ws2[f'B{row}'] = 5
ws2[f'C{row}'] = 'ft'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Initial Void Ratio (e0):'
ws2[f'B{row}'] = 0.8
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Compression Index (Cc):'
ws2[f'B{row}'] = 0.25
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Recompression Index (Cr):'
ws2[f'B{row}'] = 0.05
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'OCR:'
ws2[f'B{row}'] = 2.0
ws2[f'C{row}'] = 'Overconsolidation Ratio'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Coefficient of Consolidation (cv):'
ws2[f'B{row}'] = 0.8
ws2[f'C{row}'] = 'ft²/year'
ws2[f'B{row}'].fill = input_fill

row += 2
# Drainage Conditions
ws2[f'A{row}'] = 'DRAINAGE CONDITIONS'
ws2[f'A{row}'].font = bold_font
row += 1
ws2[f'A{row}'] = 'Drainage Condition:'
ws2[f'B{row}'] = 'Double Drainage'
ws2[f'C{row}'] = '(Top & Bottom)'
ws2[f'B{row}'].fill = input_fill

# CALCULATIONS SECTION
row += 3
ws2[f'A{row}'] = 'PRIMARY CONSOLIDATION CALCULATIONS'
ws2[f'A{row}'].font = header_font
ws2[f'A{row}'].fill = header_fill
ws2.merge_cells(f'A{row}:E{row}')

row += 2
# Current Stress State
ws2[f'A{row}'] = 'STRESS CALCULATIONS'
ws2[f'A{row}'].font = bold_font
row += 1
ws2[f'A{row}'] = 'Depth to Layer Center:'
ws2[f'B{row}'] = '=B18+B16/2'
ws2[f'C{row}'] = 'ft'
ws2[f'B{row}'].fill = calc_fill
row += 1
ws2[f'A{row}'] = 'Current Effective Stress (σ\'0):'
ws2[f'B{row}'] = '=B17*B35'
ws2[f'C{row}'] = 'psf'
ws2[f'B{row}'].fill = calc_fill
row += 1
ws2[f'A{row}'] = 'Preconsolidation Pressure (σ\'p):'
ws2[f'B{row}'] = '=B36*B22'
ws2[f'C{row}'] = 'psf'
ws2[f'B{row}'].fill = calc_fill
row += 1
ws2[f'A{row}'] = 'Applied Stress at Layer Center:'
ws2[f'B{row}'] = '=B14*0.5'
ws2[f'C{row}'] = 'psf (50% of surface stress)'
ws2[f'B{row}'].fill = calc_fill
row += 1
ws2[f'A{row}'] = 'Final Effective Stress (σ\'f):'
ws2[f'B{row}'] = '=B36+B38'
ws2[f'C{row}'] = 'psf'
ws2[f'B{row}'].fill = calc_fill

row += 2
# Settlement Calculation
ws2[f'A{row}'] = 'CONSOLIDATION SETTLEMENT'
ws2[f'A{row}'].font = bold_font
row += 1
# Check if normally or overconsolidated
ws2[f'A{row}'] = 'Stress Condition Check:'
ws2[f'B{row}'] = '=IF(B39<=B37,"Overconsolidated","Normally Consolidated")'
ws2[f'B{row}'].fill = calc_fill
row += 1
# Overconsolidated case
ws2[f'A{row}'] = 'Settlement (OC Case):'
ws2[f'B{row}'] = '=IF(B41="Overconsolidated",B16*B21/(1+B19)*LOG10(B39/B36),0)'
ws2[f'C{row}'] = 'ft (if σ\'f < σ\'p)'
ws2[f'B{row}'].fill = calc_fill
row += 1
# Normally consolidated case
ws2[f'A{row}'] = 'Settlement (NC Case):'
ws2[f'B{row}'] = '=IF(B41="Normally Consolidated",B16*(B21*LOG10(B37/B36)+B20*LOG10(B39/B37))/(1+B19),B16*(B21*LOG10(B37/B36)+B20*LOG10(B39/B37))/(1+B19))'
ws2[f'C{row}'] = 'ft (if σ\'f > σ\'p)'
ws2[f'B{row}'].fill = calc_fill
row += 1
# Total settlement
ws2[f'A{row}'] = 'Total Primary Settlement (Sp):'
ws2[f'B{row}'] = '=MAX(B42,B43)'
ws2[f'C{row}'] = 'ft'
ws2[f'B{row}'].fill = result_fill
ws2[f'B{row}'].font = bold_font
row += 1
ws2[f'A{row}'] = 'Total Primary Settlement (Sp):'
ws2[f'B{row}'] = '=B44*12'
ws2[f'C{row}'] = 'inches'
ws2[f'B{row}'].fill = result_fill
ws2[f'B{row}'].font = bold_font

row += 2
# Time-Settlement Analysis
ws2[f'A{row}'] = 'TIME-SETTLEMENT ANALYSIS'
ws2[f'A{row}'].font = bold_font
row += 1
ws2[f'A{row}'] = 'Drainage Path (H_dr):'
ws2[f'B{row}'] = '=IF(B24="Double Drainage",B16/2,B16)'
ws2[f'C{row}'] = 'ft'
ws2[f'B{row}'].fill = calc_fill
row += 1
ws2[f'A{row}'] = 'Time Factor for 50% (T50):'
ws2[f'B{row}'] = 0.197
ws2[f'C{row}'] = '(Theoretical value)'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Time Factor for 90% (T90):'
ws2[f'B{row}'] = 0.848
ws2[f'C{row}'] = '(Theoretical value)'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Time for 50% Settlement:'
ws2[f'B{row}'] = '=B48*B47^2/B23'
ws2[f'C{row}'] = 'years'
ws2[f'B{row}'].fill = calc_fill
row += 1
ws2[f'A{row}'] = 'Time for 90% Settlement:'
ws2[f'B{row}'] = '=B49*B47^2/B23'
ws2[f'C{row}'] = 'years'
ws2[f'B{row}'].fill = calc_fill

# EMBANKMENT SHEET
ws3 = embankment_sheet

# Title
ws3['A1'] = 'EMBANKMENT CONSOLIDATION ANALYSIS'
ws3['A1'].font = Font(bold=True, size=16)
ws3.merge_cells('A1:F1')

# Project Information
row = 3
ws3[f'A{row}'] = 'PROJECT INFORMATION'
ws3[f'A{row}'].font = header_font
ws3[f'A{row}'].fill = header_fill
ws3.merge_cells(f'A{row}:B{row}')

row += 1
ws3[f'A{row}'] = 'Project:'
ws3[f'B{row}'] = '[Enter Project Name]'
ws3[f'B{row}'].fill = input_fill
row += 1
ws3[f'A{row}'] = 'Engineer:'
ws3[f'B{row}'] = '[Enter Engineer Name]'
ws3[f'B{row}'].fill = input_fill
row += 1
ws3[f'A{row}'] = 'Date:'
ws3[f'B{row}'] = '[Enter Date]'
ws3[f'B{row}'].fill = input_fill

# INPUT PARAMETERS
row += 3
ws3[f'A{row}'] = 'INPUT PARAMETERS'
ws3[f'A{row}'].font = header_font
ws3[f'A{row}'].fill = header_fill
ws3.merge_cells(f'A{row}:D{row}')

row += 2
# Embankment Properties
ws3[f'A{row}'] = 'EMBANKMENT PROPERTIES'
ws3[f'A{row}'].font = bold_font
row += 1
ws3[f'A{row}'] = 'Embankment Height (H_emb):'
ws3[f'B{row}'] = 15
ws3[f'C{row}'] = 'ft'
ws3[f'B{row}'].fill = input_fill
row += 1
ws3[f'A{row}'] = 'Embankment Width (Top):'
ws3[f'B{row}'] = 50
ws3[f'C{row}'] = 'ft'
ws3[f'B{row}'].fill = input_fill
row += 1
ws3[f'A{row}'] = 'Side Slopes:'
ws3[f'B{row}'] = 2
ws3[f'C{row}'] = 'H:1V'
ws3[f'B{row}'].fill = input_fill
row += 1
ws3[f'A{row}'] = 'Embankment Unit Weight:'
ws3[f'B{row}'] = 125
ws3[f'C{row}'] = 'pcf'
ws3[f'B{row}'].fill = input_fill
row += 1
ws3[f'A{row}'] = 'Embankment Width (Bottom):'
ws3[f'B{row}'] = '=B13+2*B15*B12'
ws3[f'C{row}'] = 'ft'
ws3[f'B{row}'].fill = calc_fill

row += 2
# Foundation Soil Properties
ws3[f'A{row}'] = 'FOUNDATION SOIL PROPERTIES'
ws3[f'A{row}'].font = bold_font
row += 1
ws3[f'A{row}'] = 'Consolidating Layer Thickness:'
ws3[f'B{row}'] = 20
ws3[f'C{row}'] = 'ft'
ws3[f'B{row}'].fill = input_fill
row += 1
ws3[f'A{row}'] = 'Depth to Layer Top:'
ws3[f'B{row}'] = 5
ws3[f'C{row}'] = 'ft'
ws3[f'B{row}'].fill = input_fill
row += 1
ws3[f'A{row}'] = 'Unit Weight:'
ws3[f'B{row}'] = 115
ws3[f'C{row}'] = 'pcf'
ws3[f'B{row}'].fill = input_fill
row += 1
ws3[f'A{row}'] = 'Initial Void Ratio (e0):'
ws3[f'B{row}'] = 0.9
ws3[f'B{row}'].fill = input_fill
row += 1
ws3[f'A{row}'] = 'Compression Index (Cc):'
ws3[f'B{row}'] = 0.35
ws3[f'B{row}'].fill = input_fill
row += 1
ws3[f'A{row}'] = 'OCR:'
ws3[f'B{row}'] = 1.5
ws3[f'B{row}'].fill = input_fill
row += 1
ws3[f'A{row}'] = 'Coefficient of Consolidation:'
ws3[f'B{row}'] = 0.5
ws3[f'C{row}'] = 'ft²/year'
ws3[f'B{row}'].fill = input_fill

# CALCULATIONS SECTION
row += 3
ws3[f'A{row}'] = 'EMBANKMENT CONSOLIDATION CALCULATIONS'
ws3[f'A{row}'].font = header_font
ws3[f'A{row}'].fill = header_fill
ws3.merge_cells(f'A{row}:E{row}')

row += 2
# Stress Distribution
ws3[f'A{row}'] = 'STRESS DISTRIBUTION'
ws3[f'A{row}'].font = bold_font
row += 1
ws3[f'A{row}'] = 'Embankment Load (q_emb):'
ws3[f'B{row}'] = '=B16*B12'
ws3[f'C{row}'] = 'psf'
ws3[f'B{row}'].fill = calc_fill
row += 1
ws3[f'A{row}'] = 'Depth to Layer Center:'
ws3[f'B{row}'] = '=B19+B18/2'
ws3[f'C{row}'] = 'ft'
ws3[f'B{row}'].fill = calc_fill
row += 1
ws3[f'A{row}'] = 'Current Effective Stress:'
ws3[f'B{row}'] = '=B21*B33'
ws3[f'C{row}'] = 'psf'
ws3[f'B{row}'].fill = calc_fill
row += 1
ws3[f'A{row}'] = 'Preconsolidation Pressure:'
ws3[f'B{row}'] = '=B35*B25'
ws3[f'C{row}'] = 'psf'
ws3[f'B{row}'].fill = calc_fill
row += 1
ws3[f'A{row}'] = 'Stress Increase (Center):'
ws3[f'B{row}'] = '=B31*0.7'
ws3[f'C{row}'] = 'psf (70% at center)'
ws3[f'B{row}'].fill = calc_fill
row += 1
ws3[f'A{row}'] = 'Final Effective Stress:'
ws3[f'B{row}'] = '=B35+B37'
ws3[f'C{row}'] = 'psf'
ws3[f'B{row}'].fill = calc_fill

row += 2
# Settlement Calculation
ws3[f'A{row}'] = 'EMBANKMENT SETTLEMENT'
ws3[f'A{row}'].font = bold_font
row += 1
ws3[f'A{row}'] = 'Primary Settlement:'
ws3[f'B{row}'] = '=B18*B24/(1+B23)*LOG10(B38/B35)'
ws3[f'C{row}'] = 'ft'
ws3[f'B{row}'].fill = result_fill
ws3[f'B{row}'].font = bold_font
row += 1
ws3[f'A{row}'] = 'Primary Settlement:'
ws3[f'B{row}'] = '=B40*12'
ws3[f'C{row}'] = 'inches'
ws3[f'B{row}'].fill = result_fill
ws3[f'B{row}'].font = bold_font

row += 2
# Time Analysis
ws3[f'A{row}'] = 'TIME ANALYSIS'
ws3[f'A{row}'].font = bold_font
row += 1
ws3[f'A{row}'] = 'Drainage Path:'
ws3[f'B{row}'] = '=B18'
ws3[f'C{row}'] = 'ft (single drainage)'
ws3[f'B{row}'].fill = calc_fill
row += 1
ws3[f'A{row}'] = 'Time for 90% Settlement:'
ws3[f'B{row}'] = '=0.848*B43^2/B26'
ws3[f'C{row}'] = 'years'
ws3[f'B{row}'].fill = calc_fill

row += 2
# Immediate Settlement
ws3[f'A{row}'] = 'IMMEDIATE SETTLEMENT'
ws3[f'A{row}'].font = bold_font
row += 1
ws3[f'A{row}'] = 'Modulus of Elasticity (Est.):'
ws3[f'B{row}'] = 8000
ws3[f'C{row}'] = 'psf'
ws3[f'B{row}'].fill = input_fill
row += 1
ws3[f'A{row}'] = 'Poisson Ratio:'
ws3[f'B{row}'] = 0.4
ws3[f'B{row}'].fill = input_fill
row += 1
ws3[f'A{row}'] = 'Immediate Settlement (Est.):'
ws3[f'B{row}'] = '=B31*B17*(1-B48^2)/B47*12'
ws3[f'C{row}'] = 'inches'
ws3[f'B{row}'].fill = calc_fill

row += 2
# Total Settlement
ws3[f'A{row}'] = 'TOTAL SETTLEMENT SUMMARY'
ws3[f'A{row}'].font = bold_font
row += 1
ws3[f'A{row}'] = 'Immediate Settlement:'
ws3[f'B{row}'] = '=B49'
ws3[f'C{row}'] = 'inches'
ws3[f'B{row}'].fill = result_fill
row += 1
ws3[f'A{row}'] = 'Primary Consolidation:'
ws3[f'B{row}'] = '=B41'
ws3[f'C{row}'] = 'inches'
ws3[f'B{row}'].fill = result_fill
row += 1
ws3[f'A{row}'] = 'Total Settlement:'
ws3[f'B{row}'] = '=B51+B52'
ws3[f'C{row}'] = 'inches'
ws3[f'B{row}'].fill = result_fill
ws3[f'B{row}'].font = bold_font

# Set column widths for all sheets
for sheet in [ws1, ws2, ws3]:
    sheet.column_dimensions['A'].width = 35
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15

# Save the workbook
wb.save('Consolidation_Settlement_Analysis.xlsx')

print("Excel file 'Consolidation_Settlement_Analysis.xlsx' created successfully!")
print("Features included:")
print("- Immediate Settlement Analysis (Elastic Theory)")
print("- Primary Consolidation Analysis (Terzaghi Theory)")
print("- Embankment Loading Analysis")
print("- Time-Settlement Relationships")
print("- Alternative methods (Schmertmann SPT-based)")
print("- English units throughout (psf, pcf, ft, in)")
print("- Color-coded input/calculation/result sections")
print("- Overconsolidated and normally consolidated conditions")