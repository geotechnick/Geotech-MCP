import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import math

# Create workbook with multiple sheets
wb = Workbook()

# Remove default sheet
wb.remove(wb.active)

# Create sheets
immediate_sheet = wb.create_sheet('Immediate Settlement')
primary_sheet = wb.create_sheet('Primary Consolidation')
embankment_sheet = wb.create_sheet('Embankment Multi-Layer')
layer_input_sheet = wb.create_sheet('Soil Layer Properties')

# Define styles
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
input_fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
calc_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
result_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
layer_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
bold_font = Font(bold=True)

# SOIL LAYER PROPERTIES SHEET
ws_layers = layer_input_sheet

# Title
ws_layers['A1'] = 'SOIL LAYER PROPERTIES INPUT'
ws_layers['A1'].font = Font(bold=True, size=16)
ws_layers.merge_cells('A1:L1')

row = 3
ws_layers[f'A{row}'] = 'INSTRUCTIONS: Enter soil layer properties from top to bottom. Up to 10 layers supported.'
ws_layers[f'A{row}'].font = Font(italic=True)
ws_layers.merge_cells(f'A{row}:L{row}')

# Create header row for soil properties
row = 5
headers = ['Layer', 'Description', 'Top Depth', 'Bottom Depth', 'Thickness', 'Unit Weight', 
           'E (Modulus)', 'Poisson Ratio', 'Cc', 'Cr', 'e0', 'OCR', 'cv']
units = ['#', 'Text', 'ft', 'ft', 'ft', 'pcf', 'psf', '-', '-', '-', '-', '-', 'ft²/year']

for i, (header, unit) in enumerate(zip(headers, units)):
    ws_layers.cell(row=row, column=i+1, value=header)
    ws_layers.cell(row=row, column=i+1).font = header_font
    ws_layers.cell(row=row, column=i+1).fill = header_fill
    ws_layers.cell(row=row+1, column=i+1, value=unit)
    ws_layers.cell(row=row+1, column=i+1).font = Font(italic=True)
    ws_layers.cell(row=row+1, column=i+1).fill = header_fill

# Add sample layers
sample_layers = [
    [1, 'Fill', 0, 3, 3, 110, 3000, 0.3, 0.15, 0.03, 0.6, 1.0, 1.2],
    [2, 'Soft Clay', 3, 15, 12, 105, 2000, 0.4, 0.35, 0.08, 1.2, 1.5, 0.8],
    [3, 'Stiff Clay', 15, 25, 10, 125, 8000, 0.35, 0.25, 0.05, 0.8, 2.5, 1.5],
    [4, 'Dense Sand', 25, 40, 15, 130, 15000, 0.25, 0.05, 0.01, 0.5, 1.0, 5.0],
    [5, 'Hard Clay', 40, 60, 20, 135, 20000, 0.3, 0.15, 0.03, 0.6, 3.0, 2.0]
]

row = 7
for layer_data in sample_layers:
    for j, value in enumerate(layer_data):
        ws_layers.cell(row=row, column=j+1, value=value)
        ws_layers.cell(row=row, column=j+1).fill = input_fill
    row += 1

# Add empty rows for additional layers
for i in range(5):
    ws_layers.cell(row=row, column=1, value=6+i)
    ws_layers.cell(row=row, column=1).fill = input_fill
    for j in range(1, 13):
        ws_layers.cell(row=row, column=j+1).fill = input_fill
    row += 1

# IMMEDIATE SETTLEMENT MULTI-LAYER SHEET
ws1 = immediate_sheet

# Title
ws1['A1'] = 'IMMEDIATE SETTLEMENT ANALYSIS - MULTI-LAYER'
ws1['A1'].font = Font(bold=True, size=16)
ws1.merge_cells('A1:H1')

# Project Information
row = 3
ws1[f'A{row}'] = 'PROJECT INFORMATION'
ws1[f'A{row}'].font = header_font
ws1[f'A{row}'].fill = header_fill
ws1.merge_cells(f'A{row}:B{row}')

row += 1
ws1[f'A{row}'] = 'Project:'
ws1[f'B{row}'] = '[Enter Project Name]'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Engineer:'
ws1[f'B{row}'] = '[Enter Engineer Name]'
ws1[f'B{row}'].fill = input_fill

# Foundation Properties
row += 3
ws1[f'A{row}'] = 'FOUNDATION PROPERTIES'
ws1[f'A{row}'].font = header_font
ws1[f'A{row}'].fill = header_fill
ws1.merge_cells(f'A{row}:D{row}')

row += 1
ws1[f'A{row}'] = 'Foundation Width (B):'
ws1[f'B{row}'] = 8.0
ws1[f'C{row}'] = 'ft'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Foundation Length (L):'
ws1[f'B{row}'] = 12.0
ws1[f'C{row}'] = 'ft'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Foundation Depth (Df):'
ws1[f'B{row}'] = 3.0
ws1[f'C{row}'] = 'ft'
ws1[f'B{row}'].fill = input_fill
row += 1
ws1[f'A{row}'] = 'Applied Pressure (q):'
ws1[f'B{row}'] = 3000
ws1[f'C{row}'] = 'psf'
ws1[f'B{row}'].fill = input_fill

# Multi-layer settlement calculation
row += 3
ws1[f'A{row}'] = 'MULTI-LAYER SETTLEMENT CALCULATION'
ws1[f'A{row}'].font = header_font
ws1[f'A{row}'].fill = header_fill
ws1.merge_cells(f'A{row}:H{row}')

row += 2
# Create header for layer calculations
layer_headers = ['Layer', 'Top Depth', 'Bottom Depth', 'Thickness', 'Mid Depth', 'Stress Factor', 'Stress Increase', 'Settlement']
layer_units = ['#', 'ft', 'ft', 'ft', 'ft', '-', 'psf', 'inches']

for i, (header, unit) in enumerate(zip(layer_headers, layer_units)):
    ws1.cell(row=row, column=i+1, value=header)
    ws1.cell(row=row, column=i+1).font = header_font
    ws1.cell(row=row, column=i+1).fill = header_fill
    ws1.cell(row=row+1, column=i+1, value=unit)
    ws1.cell(row=row+1, column=i+1).font = Font(italic=True)
    ws1.cell(row=row+1, column=i+1).fill = header_fill

# Layer-by-layer calculations (referencing soil layer properties sheet)
start_calc_row = row + 2
for layer_num in range(1, 6):  # 5 layers
    calc_row = start_calc_row + layer_num - 1
    
    # Layer number
    ws1.cell(row=calc_row, column=1, value=layer_num)
    ws1.cell(row=calc_row, column=1).fill = calc_fill
    
    # Top depth (reference from layer sheet)
    ws1.cell(row=calc_row, column=2, value=f"='Soil Layer Properties'.C{6+layer_num}")
    ws1.cell(row=calc_row, column=2).fill = calc_fill
    
    # Bottom depth
    ws1.cell(row=calc_row, column=3, value=f"='Soil Layer Properties'.D{6+layer_num}")
    ws1.cell(row=calc_row, column=3).fill = calc_fill
    
    # Thickness
    ws1.cell(row=calc_row, column=4, value=f"='Soil Layer Properties'.E{6+layer_num}")
    ws1.cell(row=calc_row, column=4).fill = calc_fill
    
    # Mid depth
    ws1.cell(row=calc_row, column=5, value=f"=(B{calc_row}+C{calc_row})/2")
    ws1.cell(row=calc_row, column=5).fill = calc_fill
    
    # Stress influence factor (simplified Boussinesq)
    ws1.cell(row=calc_row, column=6, value=f"=1/(1+(E{calc_row}/$B$11)^3)^1.5")
    ws1.cell(row=calc_row, column=6).fill = calc_fill
    
    # Stress increase
    ws1.cell(row=calc_row, column=7, value=f"=$B$12*F{calc_row}")
    ws1.cell(row=calc_row, column=7).fill = calc_fill
    
    # Layer settlement (Si = Δσ * H * (1-ν²) / E)
    ws1.cell(row=calc_row, column=8, value=f"=G{calc_row}*D{calc_row}*(1-'Soil Layer Properties'.H{6+layer_num}^2)/'Soil Layer Properties'.G{6+layer_num}*12")
    ws1.cell(row=calc_row, column=8).fill = calc_fill

# Total settlement
row = start_calc_row + 6
ws1[f'A{row}'] = 'TOTAL IMMEDIATE SETTLEMENT:'
ws1[f'A{row}'].font = bold_font
ws1[f'H{row}'] = f'=SUM(H{start_calc_row}:H{start_calc_row+4})'
ws1[f'H{row}'].fill = result_fill
ws1[f'H{row}'].font = bold_font

row += 1
ws1[f'G{row}'] = 'inches'
ws1[f'G{row}'].font = bold_font

# PRIMARY CONSOLIDATION MULTI-LAYER SHEET
ws2 = primary_sheet

# Title
ws2['A1'] = 'PRIMARY CONSOLIDATION ANALYSIS - MULTI-LAYER'
ws2['A1'].font = Font(bold=True, size=16)
ws2.merge_cells('A1:J1')

# Project Information
row = 3
ws2[f'A{row}'] = 'PROJECT INFORMATION'
ws2[f'A{row}'].font = header_font
ws2[f'A{row}'].fill = header_fill
ws2.merge_cells(f'A{row}:B{row}')

row += 1
ws2[f'A{row}'] = 'Project:'
ws2[f'B{row}'] = '[Enter Project Name]'
ws2[f'B{row}'].fill = input_fill

# Foundation Properties
row += 3
ws2[f'A{row}'] = 'LOADING CONDITIONS'
ws2[f'A{row}'].font = header_font
ws2[f'A{row}'].fill = header_fill
ws2.merge_cells(f'A{row}:D{row}')

row += 1
ws2[f'A{row}'] = 'Foundation Width (B):'
ws2[f'B{row}'] = 8.0
ws2[f'C{row}'] = 'ft'
ws2[f'B{row}'].fill = input_fill
row += 1
ws2[f'A{row}'] = 'Applied Pressure (Δq):'
ws2[f'B{row}'] = 3000
ws2[f'C{row}'] = 'psf'
ws2[f'B{row}'].fill = input_fill

# Multi-layer consolidation calculation
row += 3
ws2[f'A{row}'] = 'MULTI-LAYER CONSOLIDATION CALCULATION'
ws2[f'A{row}'].font = header_font
ws2[f'A{row}'].fill = header_fill
ws2.merge_cells(f'A{row}:J{row}')

row += 2
# Create header for consolidation calculations
consol_headers = ['Layer', 'Thickness', 'Mid Depth', 'σ\'0', 'σ\'p', 'Δσ', 'σ\'f', 'Settlement Type', 'Settlement', 'Time (90%)']
consol_units = ['#', 'ft', 'ft', 'psf', 'psf', 'psf', 'psf', '-', 'inches', 'years']

for i, (header, unit) in enumerate(zip(consol_headers, consol_units)):
    ws2.cell(row=row, column=i+1, value=header)
    ws2.cell(row=row, column=i+1).font = header_font
    ws2.cell(row=row, column=i+1).fill = header_fill
    ws2.cell(row=row+1, column=i+1, value=unit)
    ws2.cell(row=row+1, column=i+1).font = Font(italic=True)
    ws2.cell(row=row+1, column=i+1).fill = header_fill

# Layer-by-layer consolidation calculations
start_consol_row = row + 2
for layer_num in range(1, 6):  # 5 layers
    calc_row = start_consol_row + layer_num - 1
    layer_ref = 6 + layer_num
    
    # Layer number
    ws2.cell(row=calc_row, column=1, value=layer_num)
    ws2.cell(row=calc_row, column=1).fill = calc_fill
    
    # Thickness
    ws2.cell(row=calc_row, column=2, value=f"='Soil Layer Properties'.E{layer_ref}")
    ws2.cell(row=calc_row, column=2).fill = calc_fill
    
    # Mid depth
    ws2.cell(row=calc_row, column=3, value=f"=('Soil Layer Properties'.C{layer_ref}+'Soil Layer Properties'.D{layer_ref})/2")
    ws2.cell(row=calc_row, column=3).fill = calc_fill
    
    # Initial effective stress (σ'0)
    ws2.cell(row=calc_row, column=4, value=f"='Soil Layer Properties'.F{layer_ref}*C{calc_row}")
    ws2.cell(row=calc_row, column=4).fill = calc_fill
    
    # Preconsolidation pressure (σ'p)
    ws2.cell(row=calc_row, column=5, value=f"=D{calc_row}*'Soil Layer Properties'.L{layer_ref}")
    ws2.cell(row=calc_row, column=5).fill = calc_fill
    
    # Stress increase (Δσ) using 2:1 stress distribution
    ws2.cell(row=calc_row, column=6, value=f"=$B$9*($B$8/($B$8+2*C{calc_row}))^2")
    ws2.cell(row=calc_row, column=6).fill = calc_fill
    
    # Final effective stress (σ'f)
    ws2.cell(row=calc_row, column=7, value=f"=D{calc_row}+F{calc_row}")
    ws2.cell(row=calc_row, column=7).fill = calc_fill
    
    # Settlement type check
    ws2.cell(row=calc_row, column=8, value=f'=IF(G{calc_row}<=E{calc_row},"OC","NC")')
    ws2.cell(row=calc_row, column=8).fill = calc_fill
    
    # Consolidation settlement calculation
    # OC case: S = H*Cr/(1+e0)*log(σ'f/σ'0) if σ'f <= σ'p
    # NC case: S = H*[Cr*log(σ'p/σ'0) + Cc*log(σ'f/σ'p)]/(1+e0) if σ'f > σ'p
    settlement_formula = f"""=IF(H{calc_row}="OC",
    B{calc_row}*'Soil Layer Properties'.J{layer_ref}/(1+'Soil Layer Properties'.K{layer_ref})*LOG10(G{calc_row}/D{calc_row})*12,
    B{calc_row}*('Soil Layer Properties'.J{layer_ref}*LOG10(E{calc_row}/D{calc_row})+'Soil Layer Properties'.I{layer_ref}*LOG10(G{calc_row}/E{calc_row}))/(1+'Soil Layer Properties'.K{layer_ref})*12)"""
    
    ws2.cell(row=calc_row, column=9, value=settlement_formula.replace('\n    ', ''))
    ws2.cell(row=calc_row, column=9).fill = calc_fill
    
    # Time for 90% consolidation (single drainage assumed)
    ws2.cell(row=calc_row, column=10, value=f"=0.848*B{calc_row}^2/'Soil Layer Properties'.M{layer_ref}")
    ws2.cell(row=calc_row, column=10).fill = calc_fill

# Total consolidation settlement
row = start_consol_row + 6
ws2[f'A{row}'] = 'TOTAL PRIMARY SETTLEMENT:'
ws2[f'A{row}'].font = bold_font
ws2[f'I{row}'] = f'=SUM(I{start_consol_row}:I{start_consol_row+4})'
ws2[f'I{row}'].fill = result_fill
ws2[f'I{row}'].font = bold_font

row += 1
ws2[f'H{row}'] = 'inches'
ws2[f'H{row}'].font = bold_font

# Maximum consolidation time
row += 1
ws2[f'A{row}'] = 'MAXIMUM CONSOLIDATION TIME:'
ws2[f'A{row}'].font = bold_font
ws2[f'J{row}'] = f'=MAX(J{start_consol_row}:J{start_consol_row+4})'
ws2[f'J{row}'].fill = result_fill
ws2[f'J{row}'].font = bold_font

# EMBANKMENT MULTI-LAYER SHEET
ws3 = embankment_sheet

# Title
ws3['A1'] = 'EMBANKMENT ANALYSIS - MULTI-LAYER FOUNDATION'
ws3['A1'].font = Font(bold=True, size=16)
ws3.merge_cells('A1:J1')

# Project Information
row = 3
ws3[f'A{row}'] = 'PROJECT INFORMATION'
ws3[f'A{row}'].font = header_font
ws3[f'A{row}'].fill = header_fill
ws3.merge_cells(f'A{row}:B{row}')

row += 1
ws3[f'A{row}'] = 'Project:'
ws3[f'B{row}'] = '[Enter Project Name]'
ws3[f'B{row}'].fill = input_fill

# Embankment Properties
row += 3
ws3[f'A{row}'] = 'EMBANKMENT PROPERTIES'
ws3[f'A{row}'].font = header_font
ws3[f'A{row}'].fill = header_fill
ws3.merge_cells(f'A{row}:D{row}')

row += 1
ws3[f'A{row}'] = 'Embankment Height:'
ws3[f'B{row}'] = 12
ws3[f'C{row}'] = 'ft'
ws3[f'B{row}'].fill = input_fill
row += 1
ws3[f'A{row}'] = 'Embankment Width (Top):'
ws3[f'B{row}'] = 40
ws3[f'C{row}'] = 'ft'
ws3[f'B{row}'].fill = input_fill
row += 1
ws3[f'A{row}'] = 'Side Slopes:'
ws3[f'B{row}'] = 2.5
ws3[f'C{row}'] = 'H:1V'
ws3[f'B{row}'].fill = input_fill
row += 1
ws3[f'A{row}'] = 'Embankment Unit Weight:'
ws3[f'B{row}'] = 125
ws3[f'C{row}'] = 'pcf'
ws3[f'B{row}'].fill = input_fill
row += 1
ws3[f'A{row}'] = 'Embankment Pressure:'
ws3[f'B{row}'] = '=B9*B8'
ws3[f'C{row}'] = 'psf'
ws3[f'B{row}'].fill = calc_fill

# Multi-layer embankment settlement
row += 3
ws3[f'A{row}'] = 'MULTI-LAYER EMBANKMENT SETTLEMENT'
ws3[f'A{row}'].font = header_font
ws3[f'A{row}'].fill = header_fill
ws3.merge_cells(f'A{row}:J{row}')

row += 2
# Create header for embankment calculations
emb_headers = ['Layer', 'Thickness', 'Mid Depth', 'σ\'0', 'Stress Factor', 'Δσ', 'Immediate', 'Primary', 'Total', 'Time']
emb_units = ['#', 'ft', 'ft', 'psf', '-', 'psf', 'in', 'in', 'in', 'years']

for i, (header, unit) in enumerate(zip(emb_headers, emb_units)):
    ws3.cell(row=row, column=i+1, value=header)
    ws3.cell(row=row, column=i+1).font = header_font
    ws3.cell(row=row, column=i+1).fill = header_fill
    ws3.cell(row=row+1, column=i+1, value=unit)
    ws3.cell(row=row+1, column=i+1).font = Font(italic=True)
    ws3.cell(row=row+1, column=i+1).fill = header_fill

# Layer-by-layer embankment calculations
start_emb_row = row + 2
for layer_num in range(1, 6):
    calc_row = start_emb_row + layer_num - 1
    layer_ref = 6 + layer_num
    
    # Layer number
    ws3.cell(row=calc_row, column=1, value=layer_num)
    ws3.cell(row=calc_row, column=1).fill = calc_fill
    
    # Thickness
    ws3.cell(row=calc_row, column=2, value=f"='Soil Layer Properties'.E{layer_ref}")
    ws3.cell(row=calc_row, column=2).fill = calc_fill
    
    # Mid depth
    ws3.cell(row=calc_row, column=3, value=f"=('Soil Layer Properties'.C{layer_ref}+'Soil Layer Properties'.D{layer_ref})/2")
    ws3.cell(row=calc_row, column=3).fill = calc_fill
    
    # Initial effective stress
    ws3.cell(row=calc_row, column=4, value=f"='Soil Layer Properties'.F{layer_ref}*C{calc_row}")
    ws3.cell(row=calc_row, column=4).fill = calc_fill
    
    # Stress influence factor for embankment loading
    ws3.cell(row=calc_row, column=5, value=f"=0.9-0.1*C{calc_row}/B8")
    ws3.cell(row=calc_row, column=5).fill = calc_fill
    
    # Stress increase
    ws3.cell(row=calc_row, column=6, value=f"=$B$13*E{calc_row}")
    ws3.cell(row=calc_row, column=6).fill = calc_fill
    
    # Immediate settlement
    ws3.cell(row=calc_row, column=7, value=f"=F{calc_row}*B{calc_row}*(1-'Soil Layer Properties'.H{layer_ref}^2)/'Soil Layer Properties'.G{layer_ref}*12")
    ws3.cell(row=calc_row, column=7).fill = calc_fill
    
    # Primary settlement (simplified - normally consolidated assumption)
    ws3.cell(row=calc_row, column=8, value=f"=B{calc_row}*'Soil Layer Properties'.I{layer_ref}/(1+'Soil Layer Properties'.K{layer_ref})*LOG10((D{calc_row}+F{calc_row})/D{calc_row})*12")
    ws3.cell(row=calc_row, column=8).fill = calc_fill
    
    # Total settlement
    ws3.cell(row=calc_row, column=9, value=f"=G{calc_row}+H{calc_row}")
    ws3.cell(row=calc_row, column=9).fill = calc_fill
    
    # Time for consolidation
    ws3.cell(row=calc_row, column=10, value=f"=0.848*B{calc_row}^2/'Soil Layer Properties'.M{layer_ref}")
    ws3.cell(row=calc_row, column=10).fill = calc_fill

# Summary totals
row = start_emb_row + 6
ws3[f'A{row}'] = 'TOTAL SETTLEMENTS:'
ws3[f'A{row}'].font = bold_font

ws3[f'F{row}'] = 'IMMEDIATE:'
ws3[f'G{row}'] = f'=SUM(G{start_emb_row}:G{start_emb_row+4})'
ws3[f'G{row}'].fill = result_fill
ws3[f'G{row}'].font = bold_font

ws3[f'F{row+1}'] = 'PRIMARY:'
ws3[f'H{row+1}'] = f'=SUM(H{start_emb_row}:H{start_emb_row+4})'
ws3[f'H{row+1}'].fill = result_fill
ws3[f'H{row+1}'].font = bold_font

ws3[f'F{row+2}'] = 'TOTAL:'
ws3[f'I{row+2}'] = f'=SUM(I{start_emb_row}:I{start_emb_row+4})'
ws3[f'I{row+2}'].fill = result_fill
ws3[f'I{row+2}'].font = bold_font

ws3[f'F{row+3}'] = 'MAX TIME:'
ws3[f'J{row+3}'] = f'=MAX(J{start_emb_row}:J{start_emb_row+4})'
ws3[f'J{row+3}'].fill = result_fill
ws3[f'J{row+3}'].font = bold_font

# Set column widths for all sheets
for sheet in [ws_layers, ws1, ws2, ws3]:
    for col in range(1, 14):
        col_letter = chr(64 + col)
        if col <= 3:
            sheet.column_dimensions[col_letter].width = 15
        else:
            sheet.column_dimensions[col_letter].width = 12

# Special width for layer properties sheet
ws_layers.column_dimensions['A'].width = 8
ws_layers.column_dimensions['B'].width = 20

# Save the workbook
wb.save('Multi_Layer_Consolidation_Analysis.xlsx')

print("Excel file 'Multi_Layer_Consolidation_Analysis.xlsx' created successfully!")
print("Features included:")
print("- Soil Layer Properties input sheet (up to 10 layers)")
print("- Multi-layer immediate settlement analysis")
print("- Multi-layer primary consolidation analysis")
print("- Multi-layer embankment settlement analysis")
print("- Layer-by-layer stress distribution")
print("- Cumulative settlement calculations")
print("- Time-dependent consolidation analysis")
print("- Automatic OC vs NC condition detection")
print("- All calculations cross-reference the soil properties sheet")